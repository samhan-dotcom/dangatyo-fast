"""
단가표 이미지 → 표 추출기  [빠른 버전 / Fast Mode]

- Claude 1회 호출만 → 15~30초, 약 50~120원
- 검증·교차검증·감사·OCR 없음
- 다운로드 파일을 session_state에 캐싱 → 버튼 클릭 시 재계산 없음
"""
import base64
import io
import json
import os
from datetime import datetime
from pathlib import Path

import anthropic
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter

from templates import TEMPLATES, get_template

# .env 로드
_env_path = Path(__file__).parent / ".env"
_example_path = Path(__file__).parent / ".env.example"
if _env_path.exists():
    load_dotenv(_env_path, override=True)
elif _example_path.exists():
    load_dotenv(_example_path, override=True)

# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------
MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 16000
APP_TITLE = "⚡ 단가표 빠른 추출기"

PRICES = {
    "claude-opus-4-7":   (5.0, 25.0),
    "claude-sonnet-4-6": (3.0, 15.0),
}


def calc_cost(usage) -> float:
    in_p, out_p = PRICES.get(MODEL, (3.0, 15.0))
    base         = getattr(usage, "input_tokens", 0)
    cache_read   = getattr(usage, "cache_read_input_tokens", 0) or 0
    cache_create = getattr(usage, "cache_creation_input_tokens", 0) or 0
    output       = usage.output_tokens
    return (
        base * in_p
        + cache_create * in_p * 1.25
        + cache_read   * in_p * 0.10
        + output * out_p
    ) / 1_000_000


def _add_tokens(usage) -> None:
    """API 응답의 토큰 수를 세션 누산기에 더한다."""
    if "usage_tokens" not in st.session_state:
        st.session_state.usage_tokens = {
            "input": 0, "output": 0, "cache_read": 0, "cache_create": 0
        }
    t = st.session_state.usage_tokens
    t["input"]        += getattr(usage, "input_tokens", 0)
    t["output"]       += getattr(usage, "output_tokens", 0)
    t["cache_read"]   += getattr(usage, "cache_read_input_tokens", 0) or 0
    t["cache_create"] += getattr(usage, "cache_creation_input_tokens", 0) or 0


# ---------------------------------------------------------------------------
# 이미지 전처리
# ---------------------------------------------------------------------------
MIN_LONG_EDGE = 1500
MAX_LONG_EDGE = 2576


def preprocess_image(image_bytes: bytes) -> tuple[bytes, str, dict]:
    img = Image.open(io.BytesIO(image_bytes))
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")

    orig_size = img.size
    info = {"original_size": orig_size, "ops": []}

    long_edge = max(img.size)
    if long_edge < MIN_LONG_EDGE:
        scale = MIN_LONG_EDGE / long_edge
        img = img.resize((int(img.size[0] * scale), int(img.size[1] * scale)), Image.LANCZOS)
        info["ops"].append(f"업스케일 ×{scale:.2f}")

    long_edge = max(img.size)
    if long_edge > MAX_LONG_EDGE:
        scale = MAX_LONG_EDGE / long_edge
        img = img.resize((int(img.size[0] * scale), int(img.size[1] * scale)), Image.LANCZOS)
        info["ops"].append(f"다운스케일 ×{scale:.2f}")

    img = img.filter(ImageFilter.UnsharpMask(radius=1.5, percent=120, threshold=2))
    info["ops"].append("언샤프 마스크")

    img = ImageEnhance.Contrast(img).enhance(1.15)
    info["ops"].append("대비 +15%")

    info["final_size"] = img.size
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue(), "image/png", info


# ---------------------------------------------------------------------------
# 추출 스키마 & 프롬프트
# ---------------------------------------------------------------------------
EXTRACTION_SCHEMA = {
    "type": "object",
    "properties": {
        "doc_title": {"type": "string"},
        "tables": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "title":   {"type": "string"},
                    "headers": {"type": "array", "items": {"type": "string"}},
                    "rows":    {"type": "array",  "items": {"type": "array", "items": {"type": "string"}}},
                    "notes":   {"type": "array",  "items": {"type": "string"}},
                },
                "required": ["title", "headers", "rows", "notes"],
                "additionalProperties": False,
            },
        },
        "general_notes": {"type": "array", "items": {"type": "string"}},
        "uncertain_cells": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "table_idx": {"type": "integer"},
                    "row_idx":   {"type": "integer"},
                    "col_idx":   {"type": "integer"},
                    "reason":    {"type": "string"},
                },
                "required": ["table_idx", "row_idx", "col_idx", "reason"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["doc_title", "tables", "general_notes", "uncertain_cells"],
    "additionalProperties": False,
}

EXTRACTION_PROMPT = """이 이미지는 통신사 단가표/정책표입니다. 이미지의 모든 정보를 정확하게 추출해주세요.

**핵심 규칙:**
1. 숫자는 본 대로 적되, 확신이 낮은 셀은 `uncertain_cells`에 표시
2. 병합 셀은 풀어서 같은 값을 반복 입력
3. 단위/기호 보존: "만", "%", "원" 그대로 유지
4. 여러 표가 있으면 각각 분리해서 tables 배열에
5. 헤더는 평탄화: 상단+하단 헤더를 합쳐서 1차원 배열로
6. 모든 행의 길이는 헤더 길이와 동일하게

7과 1, 9와 5, 0과 8, 3과 8, 6과 5 특히 주의.
음수 부호(-) 절대 누락 금지."""


# ---------------------------------------------------------------------------
# Claude API 호출 (단일)
# ---------------------------------------------------------------------------
def get_client() -> anthropic.Anthropic:
    # 1) 환경변수 (.env 또는 로컬 실행)
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    # 2) Streamlit Cloud Secrets
    if not api_key:
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY")
        except Exception:
            pass
    if not api_key:
        st.error("⚠️ ANTHROPIC_API_KEY 미설정. Streamlit Cloud → App settings → Secrets에 추가하세요.")
        st.stop()
    return anthropic.Anthropic(api_key=api_key)


def extract_tables(
    image_bytes: bytes,
    media_type: str,
    template_id: str = "auto",
) -> tuple[dict, float]:
    """이미지 → 표 데이터 (1회 호출)."""
    client = get_client()
    image_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    template = get_template(template_id)
    if template:
        system_prompt = [{"type": "text", "text": template["prompt"],
                          "cache_control": {"type": "ephemeral"}}]
        user_text = (
            "위에 정의된 양식에 정확히 맞춰 이 이미지를 추출해주세요. "
            "컬럼 순서와 이름은 양식 정의와 정확히 일치해야 합니다."
        )
    else:
        system_prompt = None
        user_text = EXTRACTION_PROMPT

    create_kwargs = dict(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        thinking={"type": "adaptive"},
        output_config={
            "effort": "low",
            "format": {"type": "json_schema", "schema": EXTRACTION_SCHEMA},
        },
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {
                    "type": "base64", "media_type": media_type, "data": image_b64,
                }},
                {"type": "text", "text": user_text},
            ],
        }],
    )
    if system_prompt:
        create_kwargs["system"] = system_prompt

    with client.messages.stream(**create_kwargs) as stream:
        response = stream.get_final_message()

    _add_tokens(response.usage)
    text_blocks = [b.text for b in response.content if b.type == "text"]
    if not text_blocks:
        block_types = [b.type for b in response.content]
        raise RuntimeError(
            f"응답에 텍스트 블록이 없습니다. 블록: {block_types}, "
            f"stop_reason: {getattr(response, 'stop_reason', '?')}"
        )
    data = json.loads(text_blocks[0])
    return data, calc_cost(response.usage)


# ---------------------------------------------------------------------------
# 표 → DataFrame
# ---------------------------------------------------------------------------
def to_dataframe(table: dict) -> pd.DataFrame:
    headers = table.get("headers", [])
    rows    = table.get("rows", [])
    if not headers:
        return pd.DataFrame(rows)
    width = len(headers)
    fixed = [
        (r + [""] * (width - len(r)))[:width] if isinstance(r, list)
        else [str(r)] + [""] * (width - 1)
        for r in rows
    ]
    return pd.DataFrame(fixed, columns=headers)


def uncertain_highlight(df: pd.DataFrame, uncertain: list[dict], t_idx: int):
    """uncertain_cells를 노란색으로 강조."""
    def styler(_):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for u in uncertain:
            if u.get("table_idx") != t_idx:
                continue
            r, c = u.get("row_idx", -1), u.get("col_idx", -1)
            if 0 <= r < len(df) and 0 <= c < len(df.columns):
                styles.iat[r, c] = "background-color: #FFF2A8; color: #6B5A00; font-weight: bold"
        return styles
    return df.style.apply(styler, axis=None)


# ---------------------------------------------------------------------------
# 다운로드 빌더 (xlsx / csv / md)
# ---------------------------------------------------------------------------
THIN        = Side(border_style="thin", color="999999")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="맑은 고딕", size=10)
TITLE_FONT  = Font(name="맑은 고딕", bold=True, size=14, color="1F4E78")
NOTE_FONT   = Font(name="맑은 고딕", italic=True, size=9, color="555555")
FILL_AI     = PatternFill("solid", start_color="FFF2A8")

INVALID_SHEET_CHARS = set("\\/?*:[]")


def _sanitize(title: str, max_len: int = 28) -> str:
    cleaned = "".join("-" if c in INVALID_SHEET_CHARS else c for c in (title or ""))
    return (cleaned.strip().strip("'")[:max_len] or "표")


def build_xlsx(doc_title: str, tables_with_dfs: list,
               general_notes: list, uncertain: list = None) -> bytes:
    uncertain = uncertain or []
    wb = Workbook()
    wb.remove(wb.active)

    used = set()
    for idx, (meta, df) in enumerate(tables_with_dfs):
        raw_title  = (meta.get("title") or f"표_{idx+1}").strip()
        sheet_name = _sanitize(raw_title)
        base, n = sheet_name, 2
        while sheet_name in used:
            sheet_name = f"{base[:25]}_{n}"; n += 1
        used.add(sheet_name)
        ws = wb.create_sheet(sheet_name)

        ws.cell(row=1, column=1, value=raw_title).font = TITLE_FONT
        if df.shape[1] > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])
        ws.row_dimensions[1].height = 22

        for c, h in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=c, value=str(h))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[3].height = 30

        # uncertain 셀 위치 맵
        unc_map = {(u["row_idx"], u["col_idx"]): u
                   for u in uncertain if u.get("table_idx") == idx}

        for r_idx, row in enumerate(df.itertuples(index=False), start=4):
            df_row = r_idx - 4
            for c, v in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c, value=v)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = BORDER
                u = unc_map.get((df_row, c - 1))
                if u:
                    cell.fill = FILL_AI
                    cell.comment = Comment(f"[AI 의심] {u.get('reason','')}", "AI 검증")

        next_row = 4 + len(df) + 1
        for note in meta.get("notes", []):
            cell = ws.cell(row=next_row, column=1, value=f"※ {note}")
            cell.font = NOTE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if df.shape[1] > 1:
                ws.merge_cells(start_row=next_row, start_column=1,
                               end_row=next_row, end_column=df.shape[1])
            ws.row_dimensions[next_row].height = 24
            next_row += 1

        for c in range(1, df.shape[1] + 1):
            max_len = max(
                [len(str(df.columns[c - 1]))]
                + [len(str(v)) for v in df.iloc[:, c - 1]]
                + [10]
            )
            ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 12), 50)

    if general_notes:
        ws = wb.create_sheet("주의사항")
        ws.cell(row=1, column=1, value="전체 주의사항").font = TITLE_FONT
        for i, note in enumerate(general_notes, start=3):
            cell = ws.cell(row=i, column=1, value=f"▶ {note}")
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[i].height = 26
        ws.column_dimensions["A"].width = 110

    if not wb.sheetnames:
        wb.create_sheet("빈 문서")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(doc_title: str, tables_with_dfs: list,
              general_notes: list, uncertain: list = None) -> str:
    uncertain = uncertain or []
    unc_set = {(u["table_idx"], u["row_idx"], u["col_idx"]) for u in uncertain}

    out = io.StringIO()
    out.write(f"# {doc_title}\n\n")
    for idx, (meta, df) in enumerate(tables_with_dfs):
        out.write(f"## {meta.get('title','')}\n")
        adf = df.copy().astype(str)
        for u in uncertain:
            if u.get("table_idx") == idx:
                r, c = u["row_idx"], u["col_idx"]
                if 0 <= r < len(adf) and 0 <= c < adf.shape[1]:
                    adf.iat[r, c] = "🟡" + adf.iat[r, c]
        out.write(adf.to_csv(index=False))
        for n in meta.get("notes", []):
            out.write(f"※ {n}\n")
        out.write("\n")
    if general_notes:
        out.write("## 전체 주의사항\n")
        for n in general_notes:
            out.write(f"▶ {n}\n")
    return out.getvalue()


def build_md(doc_title: str, tables_with_dfs: list,
             general_notes: list, uncertain: list = None) -> str:
    uncertain = uncertain or []
    parts = [f"# {doc_title}", ""]
    for idx, (meta, df) in enumerate(tables_with_dfs):
        parts.append(f"## {meta.get('title','')}")
        parts.append("")
        if not df.empty:
            adf = df.copy().astype(str)
            for u in uncertain:
                if u.get("table_idx") == idx:
                    r, c = u["row_idx"], u["col_idx"]
                    if 0 <= r < len(adf) and 0 <= c < adf.shape[1]:
                        adf.iat[r, c] = f"**🟡{adf.iat[r,c]}**"
            parts.append(adf.to_markdown(index=False))
        else:
            parts.append("_(데이터 없음)_")
        parts.append("")
        for n in meta.get("notes", []):
            parts.append(f"> {n}")
        parts.append("")
    if general_notes:
        parts.append("## 전체 주의사항")
        for n in general_notes:
            parts.append(f"- {n}")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
st.set_page_config(page_title=APP_TITLE, page_icon="⚡", layout="wide")
st.title(APP_TITLE)
st.caption("이미지 업로드 → 1회 AI 추출 → 즉시 다운로드  |  빠르고 저렴 (검증·감사 없음)")

with st.sidebar:
    st.markdown("### 📖 사용 방법")
    st.markdown("""
**① 이미지 업로드**
단가표 사진 또는 스크린샷을 올려주세요.
*(PDF → 스크린샷이 사진보다 정확합니다)*

**② 양식 선택**
통신사 양식을 선택하거나 자동감지를 사용하세요.

**③ ⚡ 추출 시작 클릭**
약 10~20초 소요됩니다.

**④ 결과 확인**
🟡 노란 셀 = AI가 불확실한 항목
→ 반드시 원본 이미지와 대조하세요.

**⑤ 수정 후 다운로드**
셀 더블클릭으로 직접 수정 가능
→ Excel / CSV / Markdown 저장
    """)
    st.divider()

    st.markdown("### ⚡ 이 버전 특징")
    st.markdown("""
- **1회 Claude 호출** (검증/감사 없음)
- **약 10~20초**, 50~120원/장
- AI가 의심하는 셀은 🟡 노란색 표시
- 빠른 확인 용도에 적합
    """)
    st.divider()

    st.markdown("### 📋 등록된 양식")
    for tid, t in TEMPLATES.items():
        st.caption(t["display_name"])
    st.divider()

    st.markdown("### 💳 API 크레딧")
    st.link_button(
        "🔗 잔여 크레딧 확인",
        "https://console.anthropic.com/settings/billing",
        use_container_width=True,
    )
    st.caption("Anthropic API는 잔액 직접 조회를 지원하지 않습니다. 위 버튼으로 Console을 바로 열 수 있습니다.")

uploaded = st.file_uploader(
    "단가표 이미지 업로드",
    type=["png", "jpg", "jpeg", "webp"],
    help="선명한 이미지일수록 정확도가 높습니다.",
)

if uploaded:
    image_bytes = uploaded.read()
    media_type  = uploaded.type or "image/png"

    if "extracted" not in st.session_state or st.session_state.get("filename") != uploaded.name:
        # 양식 선택 + 추출 버튼
        template_id = st.selectbox(
            "📋 양식 선택",
            options=list(TEMPLATES.keys()),
            format_func=lambda k: TEMPLATES[k]["display_name"],
        )
        if st.button("⚡ 추출 시작", type="primary", use_container_width=True):
            # 토큰 누산기 초기화
            st.session_state.usage_tokens = {
                "input": 0, "output": 0, "cache_read": 0, "cache_create": 0
            }
            with st.spinner("AI 분석 중... (10~20초)"):
                try:
                    start_time = datetime.now()
                    processed_bytes, processed_mime, prep_info = preprocess_image(image_bytes)
                    data, cost = extract_tables(processed_bytes, processed_mime, template_id)
                    end_time = datetime.now()
                    elapsed = (end_time - start_time).total_seconds()
                    st.session_state.update(
                        extracted      = data,
                        cost           = cost,
                        filename       = uploaded.name,
                        image_bytes    = image_bytes,
                        template_id    = template_id,
                        prep_info      = prep_info,
                        start_time     = start_time,
                        end_time       = end_time,
                        elapsed        = elapsed,
                        # 다운로드 캐시 초기화
                        dl_xlsx        = None,
                        dl_csv         = None,
                        dl_md          = None,
                        dl_built_key   = None,
                    )
                    st.rerun()
                except json.JSONDecodeError as e:
                    st.error(f"AI 응답 파싱 실패: {e}")
                except anthropic.APIError as e:
                    st.error(f"API 오류: {getattr(e, 'message', e)}")
                except Exception as e:
                    st.error(f"오류: {e}")
        else:
            st.image(image_bytes, caption="업로드된 이미지", use_container_width=True)

    else:
        # ── 결과 화면 ───────────────────────────────────────────────
        data         = st.session_state.extracted
        uncertain    = data.get("uncertain_cells", [])
        image_bytes  = st.session_state.image_bytes

        _elapsed  = st.session_state.get("elapsed", 0)
        _t_start  = st.session_state.get("start_time")
        _t_end    = st.session_state.get("end_time")

        st.success(
            f"✅ 추출 완료!  표 {len(data.get('tables',[]))}개 · "
            f"비용: 약 ${st.session_state.cost:.4f} · "
            f"소요: {_elapsed:.1f}초"
        )

        if _t_start and _t_end:
            st.caption(
                f"🕐 시작: {_t_start.strftime('%H:%M:%S')}  →  "
                f"완료: {_t_end.strftime('%H:%M:%S')}  "
                f"({_elapsed:.1f}초)"
            )

        # 토큰 사용량 상세
        _tok = st.session_state.get("usage_tokens", {})
        if _tok:
            _in   = _tok.get("input", 0)
            _out  = _tok.get("output", 0)
            _cr   = _tok.get("cache_read", 0)
            _cc   = _tok.get("cache_create", 0)
            _krw  = int(st.session_state.cost * 1_400)
            _parts = [f"입력 {_in:,}", f"출력 {_out:,}"]
            if _cc: _parts.append(f"캐시쓰기 {_cc:,}")
            if _cr: _parts.append(f"캐시읽기 {_cr:,}")
            st.caption(f"🔢 토큰: {' / '.join(_parts)} · 약 {_krw}원")

        prep = st.session_state.get("prep_info")
        if prep:
            ow, oh = prep["original_size"]
            fw, fh = prep["final_size"]
            st.caption(f"🛠️ 전처리: {ow}×{oh} → {fw}×{fh}  ({' · '.join(prep['ops'])})")

        if uncertain:
            st.info(f"🟡 AI가 {len(uncertain)}개 셀에서 확신 낮음 — 노란색 표시. 원본과 대조하세요.")

        # ── 다운로드 (상단) ───────────────────────────────────────────
        # 최초 렌더: 원본 추출 데이터로 즉시 빌드
        if st.session_state.get("dl_xlsx") is None:
            _init_tbl = [
                ({"title": t.get("title", ""), "notes": t.get("notes", [])}, to_dataframe(t))
                for t in data.get("tables", [])
            ]
            _init_notes = data.get("general_notes", [])
            st.session_state.dl_xlsx = build_xlsx("단가표", _init_tbl, _init_notes, uncertain)
            st.session_state.dl_csv  = build_csv("단가표", _init_tbl, _init_notes, uncertain).encode("utf-8-sig")
            st.session_state.dl_md   = build_md("단가표", _init_tbl, _init_notes, uncertain).encode("utf-8")
            st.session_state["_dl_doc_title"] = "단가표"

        _dl_doc  = st.session_state.get("_dl_doc_title", "단가표")
        _dl_safe = ("".join(c for c in _dl_doc if c.isalnum() or c in "_-가-힣 ").strip() or "단가표")
        _dl_ts   = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        st.markdown("### 📥 다운로드")
        st.caption(f"📁 `{_dl_safe}_{_dl_ts}`  ·  표 수정 시 자동 반영")
        _d1, _d2, _d3, _d4 = st.columns(4)
        with _d1:
            st.download_button("📊 Excel (.xlsx)", data=st.session_state.dl_xlsx,
                               file_name=f"{_dl_safe}_{_dl_ts}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with _d2:
            st.download_button("📄 CSV (.csv)", data=st.session_state.dl_csv,
                               file_name=f"{_dl_safe}_{_dl_ts}.csv",
                               mime="text/csv", use_container_width=True)
        with _d3:
            st.download_button("📝 Markdown (.md)", data=st.session_state.dl_md,
                               file_name=f"{_dl_safe}_{_dl_ts}.md",
                               mime="text/markdown", use_container_width=True)
        with _d4:
            if st.button("🔄 새 이미지", use_container_width=True):
                for k in ["extracted", "cost", "filename", "image_bytes",
                          "template_id", "prep_info", "start_time", "end_time", "elapsed",
                          "dl_xlsx", "dl_csv", "dl_md", "dl_built_key", "_dl_doc_title"]:
                    st.session_state.pop(k, None)
                st.rerun()
        st.markdown("---")

        # ── 레이아웃: 원본 | 표 ──────────────────────────────────────
        col_img, col_tbl = st.columns([1, 1.4])

        with col_img:
            st.markdown("### 📷 원본 이미지")
            st.image(image_bytes, use_container_width=True)
            st.markdown("---")
            doc_title = st.text_input("문서 제목", value="단가표")
            st.markdown("### 📝 전체 주의사항")
            gn_text = st.text_area(
                "한 줄에 하나씩",
                value="\n".join(data.get("general_notes", [])),
                height=180,
                label_visibility="collapsed",
            )
            general_notes = [ln.strip() for ln in gn_text.split("\n") if ln.strip()]

        with col_tbl:
            st.markdown("### ✏️ 추출된 표 (직접 수정 가능)")
            if uncertain:
                st.caption("🟡 노란색 = AI가 확신 낮은 셀. 원본 이미지와 대조 후 수정하세요.")
            else:
                st.caption("셀을 더블클릭해서 수정할 수 있습니다.")

            tables_with_dfs = []
            for idx, table in enumerate(data.get("tables", [])):
                unc_here = [u for u in uncertain if u.get("table_idx") == idx]
                suffix   = f"  🟡 {len(unc_here)}개 확인 필요" if unc_here else ""
                with st.expander(f"📊 {table.get('title', f'표 {idx+1}')}{suffix}", expanded=True):
                    new_title = st.text_input("제목", value=table.get("title", ""), key=f"title_{idx}")
                    df = to_dataframe(table)
                    df_show = uncertain_highlight(df, uncertain, idx) if unc_here else df
                    edited_df = st.data_editor(
                        df_show, num_rows="dynamic",
                        use_container_width=True, key=f"editor_{idx}",
                    )
                    notes_text = st.text_area(
                        "비고", value="\n".join(table.get("notes", [])),
                        height=70, key=f"notes_{idx}",
                    )
                    notes = [ln.strip() for ln in notes_text.split("\n") if ln.strip()]
                    tables_with_dfs.append(({"title": new_title, "notes": notes}, edited_df))

        # ── 다운로드 파일 재빌드 (편집 내용 반영 → 다음 렌더에 상단 버튼에 적용) ──
        build_key = (
            doc_title,
            tuple(gn_text.split("\n")),
            tuple(
                (st.session_state.get(f"title_{i}", ""),
                 st.session_state.get(f"notes_{i}", ""))
                for i in range(len(data.get("tables", [])))
            ),
        )
        if st.session_state.get("dl_built_key") != build_key:
            st.session_state.dl_xlsx      = build_xlsx(doc_title, tables_with_dfs, general_notes, uncertain)
            st.session_state.dl_csv       = build_csv(doc_title, tables_with_dfs, general_notes, uncertain).encode("utf-8-sig")
            st.session_state.dl_md        = build_md(doc_title, tables_with_dfs, general_notes, uncertain).encode("utf-8")
            st.session_state.dl_built_key = build_key
            st.session_state["_dl_doc_title"] = doc_title

else:
    st.info("👆 위에서 단가표 이미지를 업로드해주세요.")
    with st.expander("💡 팁"):
        st.markdown("""
        - 사진보다 PDF/스크린샷이 정확합니다
        - 한 번에 한 페이지씩 올리세요
        - 정밀 검증이 필요하면 **정밀 버전(app.py)**을 사용하세요
        """)
